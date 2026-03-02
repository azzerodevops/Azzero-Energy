"""Generate Excel (XLSX) reports using openpyxl.

Creates a multi-sheet workbook with:
1. Riepilogo — Overview with KPIs
2. Consumi — Energy demands breakdown
3. Risultati Tecnologie — Technology results
4. Analisi Finanziaria — Financial analysis
"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import ReportData

# Brand colours
PRIMARY_FILL = PatternFill(start_color="0097D7", end_color="0097D7", fill_type="solid")
SECONDARY_FILL = PatternFill(start_color="00B894", end_color="00B894", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
DARK_BG = PatternFill(start_color="121827", end_color="121827", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(color="0097D7", bold=True, size=14)
NORMAL_FONT = Font(color="333333", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def _style_header_row(ws, row: int, max_col: int) -> None:
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    """Auto-adjust column widths based on content."""
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 4, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def _add_riepilogo_sheet(wb: Workbook, data: ReportData) -> None:
    """Sheet 1: Overview with KPIs."""
    ws = wb.active
    ws.title = "Riepilogo"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Report — {data.analysis.name}"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Sito"
    ws["B3"] = data.analysis.site.name
    ws["A4"] = "Città"
    ws["B4"] = data.analysis.site.city or "—"
    ws["A5"] = "Anno"
    ws["B5"] = data.analysis.year
    ws["A6"] = "Scenario"
    ws["B6"] = data.scenario.name
    ws["A7"] = "Obiettivo"
    ws["B7"] = "Minimizzazione costi" if data.scenario.objective == "cost" else "Decarbonizzazione"
    ws["A8"] = "Data report"
    ws["B8"] = data.generated_at[:10]

    for row in range(3, 9):
        ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    # KPI section
    ws["A10"] = "Indicatori chiave (KPI)"
    ws["A10"].font = Font(bold=True, size=12, color="0097D7")

    kpis = [
        ("CAPEX Totale", f"€ {data.scenario.total_capex:,.0f}"),
        ("OPEX Annuo", f"€ {data.scenario.total_opex_annual:,.0f}"),
        ("Risparmio Annuo", f"€ {data.scenario.total_savings_annual:,.0f}"),
        ("Payback", f"{data.scenario.payback_years:.1f} anni" if data.scenario.payback_years else "—"),
        ("IRR", f"{data.scenario.irr:.1%}" if data.scenario.irr else "—"),
        ("NPV", f"€ {data.scenario.npv:,.0f}" if data.scenario.npv else "—"),
        ("Riduzione CO₂", f"{data.scenario.co2_reduction_percent * 100:.1f}%" if data.scenario.co2_reduction_percent else "—"),
    ]

    headers = ["Indicatore", "Valore"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=12, column=col, value=h)
    _style_header_row(ws, 12, 2)

    for i, (label, value) in enumerate(kpis, 13):
        ws.cell(row=i, column=1, value=label).border = THIN_BORDER
        ws.cell(row=i, column=2, value=value).border = THIN_BORDER

    _auto_width(ws)


def _add_consumi_sheet(wb: Workbook, data: ReportData) -> None:
    """Sheet 2: Energy demands breakdown."""
    ws = wb.create_sheet("Consumi")

    ws["A1"] = "Consumi Energetici"
    ws["A1"].font = TITLE_FONT

    headers = ["Vettore energetico", "Consumo annuo (MWh)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, len(headers))

    for i, demand in enumerate(data.demands, 4):
        ws.cell(row=i, column=1, value=demand.end_use_label).border = THIN_BORDER
        cell = ws.cell(row=i, column=2, value=demand.annual_consumption_mwh)
        cell.number_format = "#,##0.0"
        cell.border = THIN_BORDER

    # Total row
    total_row = len(data.demands) + 4
    ws.cell(row=total_row, column=1, value="TOTALE").font = Font(bold=True)
    total = sum(d.annual_consumption_mwh for d in data.demands)
    ws.cell(row=total_row, column=2, value=total).font = Font(bold=True)
    ws.cell(row=total_row, column=2).number_format = "#,##0.0"

    # Resources section
    res_start = total_row + 3
    ws.cell(row=res_start, column=1, value="Risorse Energetiche")
    ws.cell(row=res_start, column=1).font = Font(bold=True, size=12, color="0097D7")

    res_headers = ["Risorsa", "Prezzo acquisto (€/MWh)", "Fattore CO₂ (tCO₂/MWh)"]
    for col, h in enumerate(res_headers, 1):
        ws.cell(row=res_start + 2, column=col, value=h)
    _style_header_row(ws, res_start + 2, len(res_headers))

    for i, res in enumerate(data.resources, res_start + 3):
        ws.cell(row=i, column=1, value=res.resource_label).border = THIN_BORDER
        cell_price = ws.cell(row=i, column=2, value=res.buying_price)
        cell_price.number_format = "#,##0.00"
        cell_price.border = THIN_BORDER
        cell_co2 = ws.cell(row=i, column=3, value=res.co2_factor)
        cell_co2.number_format = "#,##0.000"
        cell_co2.border = THIN_BORDER

    _auto_width(ws)


def _add_tecnologie_sheet(wb: Workbook, data: ReportData) -> None:
    """Sheet 3: Technology results."""
    ws = wb.create_sheet("Risultati Tecnologie")

    ws["A1"] = "Risultati Ottimizzazione — Tecnologie"
    ws["A1"].font = TITLE_FONT

    headers = [
        "Tecnologia", "Categoria", "Capacità ottimale (kW)",
        "Produzione annua (MWh)", "CAPEX (€)", "Risparmio annuo (€)",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, len(headers))

    for i, tech in enumerate(data.scenario.tech_results, 4):
        ws.cell(row=i, column=1, value=tech.technology_name).border = THIN_BORDER
        ws.cell(row=i, column=2, value=tech.category).border = THIN_BORDER
        c = ws.cell(row=i, column=3, value=tech.optimal_capacity_kw)
        c.number_format = "#,##0.0"
        c.border = THIN_BORDER
        c = ws.cell(row=i, column=4, value=tech.annual_production_mwh)
        c.number_format = "#,##0.0"
        c.border = THIN_BORDER
        c = ws.cell(row=i, column=5, value=tech.capex)
        c.number_format = "€#,##0"
        c.border = THIN_BORDER
        c = ws.cell(row=i, column=6, value=tech.annual_savings)
        c.number_format = "€#,##0"
        c.border = THIN_BORDER

    # Totals
    total_row = len(data.scenario.tech_results) + 4
    ws.cell(row=total_row, column=1, value="TOTALE").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=sum(t.optimal_capacity_kw for t in data.scenario.tech_results)).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(t.annual_production_mwh for t in data.scenario.tech_results)).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(t.capex for t in data.scenario.tech_results)).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum(t.annual_savings for t in data.scenario.tech_results)).font = Font(bold=True)

    for col in range(3, 7):
        ws.cell(row=total_row, column=col).number_format = ws.cell(row=total_row - 1, column=col).number_format if total_row > 4 else "#,##0"

    _auto_width(ws)


def _add_finanziaria_sheet(wb: Workbook, data: ReportData) -> None:
    """Sheet 4: Financial analysis with cashflow."""
    ws = wb.create_sheet("Analisi Finanziaria")

    ws["A1"] = "Analisi Finanziaria"
    ws["A1"].font = TITLE_FONT

    # Cashflow table (simple: Year 0 = -CAPEX, Year 1..20 = savings - opex)
    headers = ["Anno", "Flusso di cassa (€)", "Cumulato (€)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    _style_header_row(ws, 3, len(headers))

    capex = data.scenario.total_capex
    annual_net = data.scenario.total_savings_annual - data.scenario.total_opex_annual

    cumulative = 0.0
    years = 20

    for year in range(years + 1):
        row = year + 4
        ws.cell(row=row, column=1, value=year).border = THIN_BORDER

        if year == 0:
            flow = -capex
        else:
            flow = annual_net

        cumulative += flow

        c = ws.cell(row=row, column=2, value=flow)
        c.number_format = "€#,##0"
        c.border = THIN_BORDER

        c = ws.cell(row=row, column=3, value=cumulative)
        c.number_format = "€#,##0"
        c.border = THIN_BORDER

    _auto_width(ws)


def generate_xlsx(data: ReportData) -> Path:
    """Generate an Excel workbook from report data.

    Returns path to the generated .xlsx file.
    """
    wb = Workbook()

    _add_riepilogo_sheet(wb, data)
    _add_consumi_sheet(wb, data)
    _add_tecnologie_sheet(wb, data)
    _add_finanziaria_sheet(wb, data)

    output_dir = Path(tempfile.mkdtemp(prefix="azzeroco2_report_"))
    filename = f"report_{data.analysis.name.replace(' ', '_')}_{data.generated_at[:10]}.xlsx"
    output_path = output_dir / filename

    wb.save(str(output_path))
    return output_path
